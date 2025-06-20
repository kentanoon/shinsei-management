# document_output_ui.py
import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
from openpyxl import load_workbook
import os
import datetime


# --- 定数 ---
DB_PATH = r"G:\マイドライブ\5-01_業務シートテンプレ\申請管理システム\application.db"
TEMPLATE_DIR = r"G:\マイドライブ\5-01_業務シートテンプレ\申請管理システム\提出書類テンプレート"  # テンプレ置き場
OUTPUT_DIR = r"G:\マイドライブ\5-01_業務シートテンプレ\申請管理システム\出力提出書類"  # 出力先

# --- 案件情報を取得 ---
def fetch_projects():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT project_code, project_name FROM projects ORDER BY id DESC")
    rows = cursor.fetchall()
    conn.close()
    return rows

def create_document(selected_project, selected_template):
    try:
        # --- ① DBから案件データ取得 ---
        project_code = selected_project.split("｜")[0]
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM projects WHERE project_code = ?", (project_code,))
        project_data = cursor.fetchone()
        conn.close()

        # --- ② テンプレート読み込み ---
        template_path = os.path.join(TEMPLATE_DIR, selected_template)
        wb = load_workbook(template_path)

        # --- ③ データを書き込み（ここでcreate_anken_gaiyoushoを呼ぶ！） ---
        if selected_template == "重要事項説明書.xlsx":
            create_anken_gaiyousho(wb, project_data)

        # --- ④ ファイル名を組み立てる ---
        today = datetime.datetime.now()
        date_str = today.strftime("%Y%m%d_%H%M")
        project_name = project_data[2]
        output_filename = f"重要事項説明書_{project_name}_{date_str}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, output_filename)

        # --- ⑤ 保存 ---
        wb.save(output_path)

        messagebox.showinfo("完了", f"書類作成が完了しました！\n\n{output_path}")

    except Exception as e:
        messagebox.showerror("エラー", str(e))


def create_anken_gaiyousho(wb, project_data):
    """
    案件概要書用のデータ流し込み関数（最新版）
    
    :param wb: openpyxlで読み込んだワークブック
    :param project_data: DBから取得した案件データ（タプル）
    """
    try:
        ws = wb.active  # 1枚目のシートを対象にする

        # --- ① C6セルに案件名を書き込み ---
        ws["E19"] = project_data[2]  # project_nameはインデックス2番
                                    

    except Exception as e:
        raise Exception(f"案件概要書作成中にエラー発生: {str(e)}")

# --- UI ---
root = tk.Tk()
root.title("資料出力")
root.geometry("600x400")

# 案件選択
tk.Label(root, text="案件を選んでください").pack(pady=10)
projects = fetch_projects()
project_var = tk.StringVar()
project_combo = ttk.Combobox(root, textvariable=project_var, width=50)
project_combo['values'] = [f"{c}｜{n}" for c, n in projects]
project_combo.pack(pady=5)


# 出力ボタン
# --- 出力ボタン（テンプレート選択せず固定ファイル使用）---
tk.Button(root, text="書類作成", font=("Arial", 12, "bold"),
          command=lambda: create_document(project_var.get(), "重要事項説明書.xlsx")).pack(pady=20)

root.mainloop()
