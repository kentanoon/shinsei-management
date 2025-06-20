# document_output_ui.py
import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
from openpyxl import load_workbook
import os
import datetime


# --- 定数 ---
DB_PATH = r"G:\マイドライブ\5-01_業務シートテンプレ\申請管理システム\application.db"
TEMPLATE_DIR = r"G:\マイドライブ\5-01_業務シートテンプレ\申請管理システム\提出書類テンプレート"  # テンプレ置き場
OUTPUT_DIR = r"G:\マイドライブ\5-01_業務シートテンプレ\申請管理システム\出力提出書類"  # 出力先

# --- 案件情報を取得 ---
def fetch_projects():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT project_code, project_name FROM projects ORDER BY id DESC")
    rows = cursor.fetchall()
    conn.close()
    return rows

def create_document(selected_project, selected_template):
    try:
        # --- ① DBから案件データ取得 ---
        project_code = selected_project.split("｜")[0]
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM projects WHERE project_code = ?", (project_code,))
        project_data = cursor.fetchone()
        conn.close()

        # --- ② テンプレート読み込み ---
        template_path = os.path.join(TEMPLATE_DIR, selected_template)
        wb = load_workbook(template_path)

        # --- ③ データを書き込み（ここでcreate_anken_gaiyoushoを呼ぶ！） ---
        if selected_template == "案件概要書.xlsx":
            create_anken_gaiyousho(wb, project_data)

        # --- ④ ファイル名を組み立てる ---
        today = datetime.datetime.now()
        date_str = today.strftime("%Y%m%d_%H%M")
        project_name = project_data[2]
        output_filename = f"案件概要書_{project_name}_{date_str}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, output_filename)

        # --- ⑤ 保存 ---
        wb.save(output_path)

        messagebox.showinfo("完了", f"書類作成が完了しました！\n\n{output_path}")

    except Exception as e:
        messagebox.showerror("エラー", str(e))


def create_anken_gaiyousho(wb, project_data):
    """
    案件概要書用のデータ流し込み関数（最新版）
    
    :param wb: openpyxlで読み込んだワークブック
    :param project_data: DBから取得した案件データ（タプル）
    """
    try:
        ws = wb.active  # 1枚目のシートを対象にする

        # --- ① C6セルに案件名を書き込み ---
        ws["C6"] = project_data[2]  # project_nameはインデックス2番
        ws["C7"] = project_data[6]  
        ws["C8"] = project_data[5]  
        ws["C9"] = project_data[15]  
        ws["F6"] = project_data[13]
        ws["F7"] = project_data[11]
        ws["F8"] = project_data[10]
        ws["C11"] = project_data[16]
        ws["E11"] = project_data[17]
        ws["G11"] = project_data[18]      
        ws["C12"] = project_data[19]
        ws["E12"] = project_data[20]
        ws["F12"] = project_data[21]
        ws["C13"] = project_data[22]
        ws["C15"] = project_data[24]
        ws["E15"] = project_data[25]
        ws["G15"] = project_data[26]
        ws["C16"] = project_data[27]
        ws["E16"] = project_data[28]
        ws["C17"] = project_data[29]
        ws["E17"] = project_data[30]
        ws["C21"] = project_data[31]
        ws["E21"] = project_data[32]
        ws["G21"] = project_data[33]                      
        ws["C22"] = project_data[34]
        ws["E22"] = project_data[35]
        ws["C24"] = project_data[36]
        ws["E24"] = project_data[37]
        ws["G24"] = project_data[38]
        ws["C25"] = project_data[39]
        ws["E25"] = project_data[40]                                        

        # --- ② 工程表を作成（ここでは仮の工程データを使う） ---
        koutei_list = [
            {"工程名": "設計", "開始日": datetime.date(2025, 5, 1), "終了日": datetime.date(2025, 5, 8)},
            {"工程名": "申請準備", "開始日": datetime.date(2025, 5, 9), "終了日": datetime.date(2025, 5, 13)},
            {"工程名": "申請", "開始日": datetime.date(2025, 5, 14), "終了日": datetime.date(2025, 5, 15)},
            {"工程名": "審査", "開始日": datetime.date(2025, 5, 16), "終了日": datetime.date(2025, 5, 27)},
            {"工程名": "訂正", "開始日": datetime.date(2025, 5, 28), "終了日": datetime.date(2025, 6, 2)},
            {"工程名": "合格", "開始日": datetime.date(2025, 6, 3), "終了日": datetime.date(2025, 6, 4)}
        ]

        # --- ③ 今日以降の工程だけをフィルタリング ---
        today = datetime.date.today()
        filtered_koutei_list = [step for step in koutei_list if step["開始日"] >= today]

        # --- ④ 工程表を書き込み ---
        start_row = 30
        start_col = 2  # B列スタート

        for idx, step in enumerate(filtered_koutei_list):
            row = start_row + idx
            ws.cell(row=row, column=start_col, value=step["工程名"])              # 工程名
            ws.cell(row=row, column=start_col + 1, value=step["開始日"])           # 開始日
            ws.cell(row=row, column=start_col + 2, value=step["終了日"])           # 終了日

        # --- ⑤ 日付の書式を整える ---
        for col in [start_col + 1, start_col + 2]:
            for idx in range(len(filtered_koutei_list)):
                ws.cell(row=start_row + idx, column=col).number_format = 'yyyy/mm/dd'

    except Exception as e:
        raise Exception(f"案件概要書作成中にエラー発生: {str(e)}")

# --- UI ---
root = tk.Tk()
root.title("資料出力")
root.geometry("600x400")

# 案件選択
tk.Label(root, text="案件を選んでください").pack(pady=10)
projects = fetch_projects()
project_var = tk.StringVar()
project_combo = ttk.Combobox(root, textvariable=project_var, width=50)
project_combo['values'] = [f"{c}｜{n}" for c, n in projects]
project_combo.pack(pady=5)


# 出力ボタン
# --- 出力ボタン（テンプレート選択せず固定ファイル使用）---
tk.Button(root, text="書類作成", font=("Arial", 12, "bold"),
          command=lambda: create_document(project_var.get(), "案件概要書.xlsx")).pack(pady=20)

root.mainloop()
