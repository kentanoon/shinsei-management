import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
from openpyxl import load_workbook
import os
import win32com.client
import csv

# --- パス設定 ---
OUTPUT_DIR = r"G:\マイドライブ\5-01_業務シートテンプレ\申請管理システム\出力見積書"
DB_PATH = r"G:\マイドライブ\5-01_業務シートテンプレ\申請管理システム\application.db"
ESTIMATE_DB_PATH = r"G:\マイドライブ\5-01_業務シートテンプレ\申請管理システム\estimate.db"
CSV_TEMPLATE_PATH = r"G:\マイドライブ\5-01_業務シートテンプレ\申請管理システム\estimate_templates.csv"
TEMPLATE_PATH = r"G:\マイドライブ\5-01_業務シートテンプレ\申請管理システム\見積テンプレート.xlsm"

# 定数
H_PAD = 1
V_PAD = 1

# --- estimates.db 初期化 ---
def init_estimate_db():
    conn = sqlite3.connect(ESTIMATE_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS estimates (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        project_code TEXT,
        estimate_code TEXT,
        created_date TEXT,
        total_price INTEGER,
        notes TEXT
    )
    """)
    conn.commit()
    conn.close()

def init_estimate_detail_db():
    conn = sqlite3.connect(ESTIMATE_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS estimate_details (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        estimate_id INTEGER,
        item_name TEXT,
        detail TEXT,
        quantity REAL,
        unit TEXT,
        unit_price INTEGER,
        amount INTEGER,
        notes TEXT,
        is_checked INTEGER DEFAULT 1,  -- ★追加！
        FOREIGN KEY (estimate_id) REFERENCES estimates(id)
    )
    """)
    conn.commit()
    conn.close()

init_estimate_db()
init_estimate_detail_db()


# --- CSVテンプレート読み込み ---
def load_templates_from_csv(filepath=CSV_TEMPLATE_PATH):
    templates = {}
    try:
        with open(filepath, newline='', encoding='utf-8-sig') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                category = row.get("申請種類", "").strip()
                if not category:
                    continue
                item = [
                    row.get("項目名", "").strip(),
                    row.get("詳細", "").strip(),
                    row.get("数量", "").strip(),
                    row.get("単位", "").strip(),
                    row.get("単価", "").strip(),
                    row.get("金額", "").strip(),
                    row.get("備考", "").strip()
                ]
                templates.setdefault(category, []).append(item)
    except Exception as e:
        messagebox.showerror("CSV読み込みエラー", str(e))
    return templates

estimate_templates = load_templates_from_csv()

# --- DBフェッチ ---
def fetch_projects():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT project_code, project_name FROM projects ORDER BY id DESC")
    rows = cursor.fetchall()
    conn.close()
    return rows

def fetch_estimate_summary(project_code):
    conn = sqlite3.connect(ESTIMATE_DB_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT estimate_code, created_date
        FROM estimates
        WHERE project_code = ?
        ORDER BY created_date DESC
        LIMIT 1
    """, (project_code,))
    result = cursor.fetchone()
    conn.close()
    return result

def simple_selection_popup(options):
    popup = tk.Toplevel(root)
    popup.title("過去の見積履歴選択")
    popup.geometry("400x300")

    listbox = tk.Listbox(popup, selectmode="single")
    for opt in options:
        listbox.insert(tk.END, opt)
    listbox.pack(fill="both", expand=True, padx=10, pady=10)

    selected_index = []

    def select_and_close():
        if listbox.curselection():
            selected_index.append(listbox.curselection()[0])
        popup.destroy()

    btn = tk.Button(popup, text="選択", command=select_and_close)
    btn.pack(pady=10)

    popup.grab_set()
    root.wait_window(popup)

    return selected_index[0] if selected_index else None



# --- 面積取得 ---
def get_project_areas(project_code):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT total_area, building_area FROM projects WHERE project_code = ?", (project_code,))
        result = cursor.fetchone()
    except Exception as e:
        messagebox.showerror("DBエラー", f"面積情報の取得に失敗しました\n{str(e)}")
        result = None
    finally:
        conn.close()
    if result:
        return result[0], result[1]
    else:
        return None, None

# --- 金額計算（カンマ付） ---
def calculate_amount(entry_qty, entry_price, entry_amount):
    try:
        qty_text = entry_qty.get().replace(",", "")
        price_text = entry_price.get().replace(",", "")
        qty = float(qty_text) if qty_text else 0
        price = float(price_text) if price_text else 0
        amount = qty * price
        entry_amount.delete(0, tk.END)
        entry_amount.insert(0, "{:,}".format(round(amount)))
    except Exception:
        entry_amount.delete(0, tk.END)
        entry_amount.insert(0, "0")

def format_number(val):
    try:
        num = float(val.replace(',', ''))
        return "{:,}".format(round(num))
    except ValueError:
        return val

def format_with_commas(entry):
    value = entry.get().replace(",", "")
    if value.isdigit():
        formatted = "{:,}".format(int(value))
        entry.delete(0, tk.END)
        entry.insert(0, formatted)

# --- 明細入力画面 ---
def proceed_to_estimate():
    global check_rows, new_win, selected_client_name, selected_project_name, selected_estimate_id

    if not selected_project_code:
        messagebox.showwarning("未選択", "案件を選択してください")
        return

    selected_apps = [k for k, v in application_vars.items() if v.get()]

    # --- 過去見積を選んでいる場合は申請内容チェックをスキップ ---
    if not selected_apps and not selected_estimate_id:
        messagebox.showwarning("未選択", "申請内容を選択してください")
        return


    rebuild_data = []

    total_area, building_area = get_project_areas(selected_project_code)

    new_win = tk.Toplevel(root)
    new_win.title("見積明細入力")
    new_win.geometry("1700x1000")

    container = tk.Frame(new_win)
    container.pack(padx=H_PAD*2, pady=V_PAD*2)

    headers = ["出力", "項目", "詳細", "数量", "単位", "単価", "金額", "備考"]
    widths = [5, 20, 65, 5, 5, 10, 10, 70]

    for col, (text, w) in enumerate(zip(headers, widths)):
        tk.Label(container, text=text, width=w, anchor="center", font=("Arial", 11)).grid(row=0, column=col, padx=H_PAD, pady=V_PAD)

    check_rows = []
    row_idx = 1

    # --- ここで復元できるか確認する！ ---
    rebuild_data = []
    if selected_estimate_id:
        conn = sqlite3.connect(ESTIMATE_DB_PATH)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT item_name, detail, quantity, unit, unit_price, amount, notes, is_checked
            FROM estimate_details
            WHERE estimate_id = ?
        """, (selected_estimate_id,))
        rebuild_data = cursor.fetchall()
        conn.close()

    if rebuild_data:
        for values in rebuild_data:
            is_checked = values[7] if len(values) > 7 else 1  # 追加列なので念のため
            var = tk.BooleanVar(value=True if is_checked else False)

            tk.Checkbutton(container, variable=var).grid(
                row=row_idx, column=0, padx=H_PAD, pady=V_PAD, sticky="w"
            )

            entries = []
            for i, val in enumerate(values[:7]):  # 先頭7つだけエントリ化
                w = widths[i + 1]
                ent = tk.Entry(container, width=w, font=("Arial", 11), justify="right" if i in (2, 4, 5) else "left")
                ent.grid(row=row_idx, column=i+1, padx=H_PAD, pady=V_PAD, sticky="w")

                if i in (2, 4, 5):
                    ent.insert(0, format_number(str(val)))
                else:
                    ent.insert(0, str(val))

                entries.append(ent)

            if len(entries) >= 6:
                entry_qty = entries[2]
                entry_price = entries[4]
                entry_amount = entries[5]
                calculate_amount(entry_qty, entry_price, entry_amount)
                format_with_commas(entry_qty)
                format_with_commas(entry_price)
                format_with_commas(entry_amount)

                entry_qty.bind("<KeyRelease>", lambda event, eq=entry_qty, ep=entry_price, ea=entry_amount: [calculate_amount(eq, ep, ea), format_with_commas(eq), format_with_commas(ea)])
                entry_price.bind("<KeyRelease>", lambda event, eq=entry_qty, ep=entry_price, ea=entry_amount: [calculate_amount(eq, ep, ea), format_with_commas(ep), format_with_commas(ea)])

            check_rows.append((var, entries))
            row_idx += 1


    else:
        # --- 新規作成（テンプレートから）
        target_items = [
            "申請図面作成", "特例無し図面作成", "計算書", "プラン作成費用",
            "基礎伏図作成", "準防火地域・防火地域", "設計監理業務",
            "壁量計算書作成", "LV計算書作成", "準防火地域・防火地域加算"
        ]

        for key in selected_apps:
            for values in estimate_templates.get(key, []):
                var = tk.BooleanVar(value=True)
                tk.Checkbutton(container, variable=var).grid(row=row_idx, column=0, padx=H_PAD, pady=V_PAD, sticky="w")

                entries = []
                item_name = values[0] if len(values) > 0 else ""

                for i in range(len(widths[1:])):
                    val = values[i] if i < len(values) else ""
                    w = widths[i + 1]
                    ent = tk.Entry(container, width=w, font=("Arial", 11), justify="right" if i in (2, 4, 5) else "left")
                    ent.grid(row=row_idx, column=i+1, padx=H_PAD, pady=V_PAD, sticky="w")

                    if i == 2 and item_name in target_items:
                        if item_name == "基礎伏図作成" and building_area:
                            val = building_area
                        elif total_area:
                            val = total_area

                    if i in (2, 4, 5):
                        ent.insert(0, format_number(val))
                    else:
                        ent.insert(0, val)

                    entries.append(ent)

                if len(entries) >= 6:
                    entry_qty = entries[2]
                    entry_price = entries[4]
                    entry_amount = entries[5]
                    calculate_amount(entry_qty, entry_price, entry_amount)
                    format_with_commas(entry_qty)
                    format_with_commas(entry_price)
                    format_with_commas(entry_amount)

                    entry_qty.bind("<KeyRelease>", lambda event, eq=entry_qty, ep=entry_price, ea=entry_amount: [calculate_amount(eq, ep, ea), format_with_commas(eq), format_with_commas(ea)])
                    entry_price.bind("<KeyRelease>", lambda event, eq=entry_qty, ep=entry_price, ea=entry_amount: [calculate_amount(eq, ep, ea), format_with_commas(ep), format_with_commas(ea)])

                check_rows.append((var, entries))
                row_idx += 1

    for c in range(len(headers)):
        container.grid_columnconfigure(c, pad=H_PAD)

    tk.Button(new_win, text="Excel→PDF出力", command=lambda: export_to_excel(new_win, selected_client_name, selected_project_name)).pack(pady=V_PAD*5)



def ask_save_method():
    popup = tk.Toplevel(root)
    popup.title("保存方法を選択")
    popup.geometry("300x150")

    result = {"choice": None}

    def choose_overwrite():
        result["choice"] = "overwrite"
        popup.destroy()

    def choose_newsave():
        result["choice"] = "new"
        popup.destroy()

    tk.Label(popup, text="保存方法を選択してください", font=("Arial", 12)).pack(pady=10)
    tk.Button(popup, text="上書き保存", command=choose_overwrite, width=15).pack(pady=5)
    tk.Button(popup, text="新規保存", command=choose_newsave, width=15).pack(pady=5)

    popup.grab_set()
    root.wait_window(popup)

    return result["choice"]

# --- Excel出力 ---
def export_to_excel(new_win, selected_client_name, selected_project_name):
    import datetime

    save_method = ask_save_method()
    if save_method is None:
        messagebox.showinfo("中断", "保存がキャンセルされました")
        new_win.destroy()
        return

    now = datetime.datetime.now()
    today = now.strftime("%Y-%m-%d %H:%M")
    timestamp = now.strftime("%Y%m%d_%H%M")
    safe_project_name = selected_project_name.strip().replace(" ", "").replace("/", "_").replace("\\", "_")
    fname = os.path.join(OUTPUT_DIR, f"見積書_{safe_project_name}_{timestamp}.xlsm")

    conn = sqlite3.connect(ESTIMATE_DB_PATH)
    cursor = conn.cursor()

    if save_method == "overwrite" and selected_estimate_id:
        estimate_id = selected_estimate_id
        cursor.execute("DELETE FROM estimate_details WHERE estimate_id = ?", (estimate_id,))
        conn.commit()
    else:
        estimate_code = f"{selected_project_code}_{now.strftime('%Y%m%d_%H%M')}"
        cursor.execute("""
            INSERT INTO estimates (project_code, estimate_code, created_date, total_price, notes)
            VALUES (?, ?, ?, ?, ?)
        """, (selected_project_code, estimate_code, today, 0, ""))
        estimate_id = cursor.lastrowid
        conn.commit()

    wb = load_workbook(TEMPLATE_PATH, keep_vba=True)
    ws = wb["明細"]
    ws_cover = wb["見積書表紙"]
    ws_cover["C9"] = selected_client_name
    ws_cover["F13"] = selected_project_name
    ws_cover["C1"] = selected_project_code

    r = 4
    total_price = 0
    for var, ents in check_rows:
        if var.get():
            is_checked = 1
        else:
            is_checked = 0

        vals = []
        db_vals = []
        for idx, e in enumerate(ents):
            value = e.get().replace(",", "")
            if idx not in (2, 4, 5):
                vals.append(value)
                db_vals.append(value)
            else:
                num_value = int(value) if value else 0
                vals.append(num_value)
                db_vals.append(num_value)
                if idx == 5:
                    total_price += num_value

        # チェックONの時だけExcel出力
        if var.get():
            ws[f"B{r}"], ws[f"C{r}"], ws[f"D{r}"], ws[f"E{r}"], ws[f"F{r}"], ws[f"G{r}"], ws[f"H{r}"] = vals
            r += 1

        # DB保存はチェック関係なく常に保存
        cursor.execute("""
            INSERT INTO estimate_details (estimate_id, item_name, detail, quantity, unit, unit_price, amount, notes, is_checked)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (estimate_id, db_vals[0], db_vals[1], db_vals[2], db_vals[3], db_vals[4], db_vals[5], db_vals[6], is_checked))

    conn.commit()
    conn.close()

    wb.save(fname)

    excel = win32com.client.Dispatch("Excel.Application")
    wb_com = excel.Workbooks.Open(os.path.abspath(fname))
    ws_com = wb_com.Worksheets("見積書表紙")
    pdf_path = fname.replace(".xlsm", ".pdf")
    ws_com.ExportAsFixedFormat(0, pdf_path)
    wb_com.Close(False)
    excel.Quit()

    messagebox.showinfo("完了", f"Excel→PDF出力が完了しました\n{pdf_path}")
    new_win.destroy()



# --- メイン画面 ---
init_estimate_db()
root = tk.Tk()
root.title("見積作成：案件・申請内容選択")
root.geometry("700x600")

project_frame = tk.LabelFrame(root, text="案件の選択", padx=H_PAD*2, pady=V_PAD*2)
project_frame.pack(fill="x", padx=H_PAD*2, pady=V_PAD*2)
project_list = fetch_projects()
project_var = tk.StringVar()
project_combo = ttk.Combobox(project_frame, textvariable=project_var, width=60)
project_combo['values'] = [f"{c}｜{n}" for c, n in project_list]
project_combo.grid(row=0, column=0, padx=H_PAD, pady=V_PAD)
summary_label = tk.Label(project_frame, text="", fg="gray")
summary_label.grid(row=1, column=0, sticky="w", padx=H_PAD)
selected_project_code = None
selected_client_name = ""
selected_project_name = ""
selected_estimate_id = None  # ★ここにちゃんと置く！


def on_project_select(e):
    global selected_project_code, selected_client_name, selected_project_name, selected_estimate_id
    sel = project_combo.get()
    if "｜" in sel:
        selected_project_code, selected_project_name = sel.split("｜")

        # 発注者名もapplication.dbから取得する！
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT client_name FROM projects WHERE project_code = ?", (selected_project_code,))
        res = cursor.fetchone()
        conn.close()

        if res:
            selected_client_name = res[0]
        else:
            selected_client_name = ""

        # --- ここから履歴取得 ---
        conn = sqlite3.connect(ESTIMATE_DB_PATH)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, created_date
            FROM estimates
            WHERE project_code = ?
            ORDER BY created_date ASC
        """, (selected_project_code,))
        histories = cursor.fetchall()
        conn.close()

        if histories:
            history_list = []
            for idx, h in enumerate(histories, start=1):
                display_text = f"{idx}. {h[1]}"  # h[1] = created_date
                history_list.append(display_text)

            selected = simple_selection_popup(history_list)
            if selected is not None:
                selected_estimate_id = histories[selected][0]
            else:
                selected_estimate_id = None
        else:
            selected_estimate_id = None

project_combo.bind("<<ComboboxSelected>>", on_project_select)

application_frame = tk.LabelFrame(root, text="申請内容選択", padx=H_PAD*2, pady=V_PAD*2)
application_frame.pack(fill="both", expand=True, padx=H_PAD*2, pady=V_PAD*2)
application_vars = {}
for i, label in enumerate(estimate_templates.keys()):
    var = tk.BooleanVar()
    cb = tk.Checkbutton(application_frame, text=label, variable=var)
    cb.grid(row=i//2, column=i%2, sticky="w", padx=H_PAD, pady=V_PAD)
    application_vars[label] = var

tk.Button(root, text="次へ", command=proceed_to_estimate).pack(pady=V_PAD*5)

root.mainloop()
